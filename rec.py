#EXCEL SHEET CODE

from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl import drawing

def output1(filename, sheet, num, frames, name, tym, rec):
    my_file = Path('recognized_data/' + filename + str(datetime.now().date()) + '.xlsx')
    if my_file.is_file():
        book = load_workbook('recognized_data/' + filename + str(datetime.now().date()) + '.xlsx')
        sh = book[sheet]
        # file exists

    else:
        book = Workbook()
        sh = book.active
        sh.title = sheet
        sh.append(('Frames', 'Name', 'Date&Time','Recognized/Not'))

    sh.column_dimensions['A'].width = 50
    sh.column_dimensions['B'].width = 50
    sh.column_dimensions['C'].width = 50
    sh.column_dimensions['D'].width = 50
    sh.row_dimensions[num].height = 200
    img = drawing.image.Image(frames)
    sh.add_image(img, 'A' + str(num))
    sh['B' + str(num)] = name
    sh['C' + str(num)] = tym
    sh['D' + str(num)] = rec

    fullname = filename + str(datetime.now().date()) + '.xlsx'
    book.save('recognized_data/' + fullname)
    return fullname